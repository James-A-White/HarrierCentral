import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---- data -----------------------------------------------------------------
# groups -> list of (function short name, SP(s), current server gate, current app/UI gate, notes)
groups = [
    ("Hash Cash", [
        ("View payment report", "getPaymentReport", "RA + admin*", "canManageHashCash", ""),
        ("Take payment (check-in)", "processPayment", "UNGATED (bug)", "none on check-in button", "P0 security gap"),
        ("Bulk payment", "processBulkPayment", "RA + admin*", "canManageRuns (snackbar)", ""),
        ("Manage receipts (expenses)", "addEditReceipt", "HashFlash|GM|VGM|RA + admin*", "canManageHashCash", ""),
    ]),
    ("Runs & Events", [
        ("Create / edit runs", "addEditEvent", "UNGATED (bug)", "canManageRuns", "P0 security gap"),
        ("Print QR codes", "(client)", "n/a", "canManageRuns", ""),
        ("Manage attendance", "setEventAttendence / setBulkEventAttendence", "HashFlash|GM|VGM|RA + admin*", "canManageRuns / none", ""),
        ("Copy RSVPs between runs", "copyEventRsvps", "HashFlash|GM|VGM|RA + admin*", "(no button audited)", ""),
        ("PackTrack trim (admin)", "(GetPositions/Delete)", "n/a", "canManageRuns | admin", ""),
    ]),
    ("Down Downs & Awards", [
        ("Award list (drinks)", "(client + downdown SPs)", "-", "canManageRuns & canManageAwards", ""),
        ("Manage Down Downs", "addDownDown / cancel / mark / update", "GM|RA (attendee to add)", "isGm|isRa; live charges UNGATED", ""),
    ]),
    ("Members", [
        ("Manage members (roster)", "syncKennelAdminData", "HashFlash|GM|VGM|RA + admin*", "canManageMembers", ""),
        ("View invite codes", "getInviteCode", "SuperAdmin|ManageMembers or self", "(no button audited)", ""),
        ("Assign app-access flags", "joinKennel (role/flag write)", "UNGATED (CRITICAL)", "isSuperAdmin", "P0 priv-escalation"),
        ("Assign mismanagement roles", "joinKennel (role/flag write)", "UNGATED (CRITICAL)", "isSuperAdmin", "P0 priv-escalation"),
    ]),
    ("Photos", [
        ("Review / approve photos", "getKennelPendingPhotos", "HashFlash|GM|VGM|RA", "isHashFlash|GM|VGM|RA", ""),
        ("Edit photo status / caption", "updatePhotoStatus / updatePhotoCaption", "HashFlash|GM|VGM|RA", "(in review page)", ""),
        ("Batch / view all photos", "batchUpdatePhotoStatus / getRunAllPhotos", "HashFlash|GM|VGM|RA|WebMeister", "isAdmin", "mask drift"),
    ]),
    ("Web / Newsletter", [
        ("Write / save Hash Trash", "saveHashTrash", "HashFlash|GM|WebMeister", "isHashFlash|isWebMeister|isGm", ""),
        ("View Hash Trash drafts", "getHashTrash", "HashFlash|GM|VGM|RA|WebMeister", "(no button audited)", "mask drift"),
    ]),
    ("Kennel", [
        ("Manage kennel settings", "(portal / editor)", "-", "isAdmin", "ManageKennel flag unused in app"),
    ]),
    ("Songs", [
        ("Manage songs", "selectSong (RSVP-gated)", "(unenforced for admin)", "(none)", "ManageSongs flag unused"),
    ]),
    ("Chat", [
        ("Post event chat message", "sendEventMessage", "UNGATED", "open to all", "P1: no membership check"),
    ]),
]

appaccess_ref = ["Super Admin (flag)", "Admin (flag)", "Hare (this run only) *"]
mm_roles = [
    "On Mismanagement", "GM (Grand Master)", "VGM (Vice GM)", "RA (Religious Advisor)",
    "Beer Meister", "Hash Flash", "On Sec", "Song Meister", "Trail Master", "Hare Raiser",
    "Hash Cash", "Scribe", "Web Meister", "Hash Hugs", "Hash Ho", "Haberdasher",
    "Hash Sweep", "Hash Trash", "Hash Bank", "Event Meister", "Communications", "Other",
]

# flat function list for columns
functions = []  # (group, short)
for gname, items in groups:
    for it in items:
        functions.append((gname, it[0]))

ALL_FN = {f[1] for f in functions}

# Best-guess pre-fill: role display name -> set of function short-names to mark "Y".
# These are DEFAULTS for review — the flag override layer is separate.
FILL = {
    "Super Admin (flag)": set(ALL_FN),
    "Admin (flag)": set(ALL_FN),
    # Hare = run-scoped: only for the run they are haring.
    "Hare (this run only) *": {"Manage receipts (expenses)", "Create / edit runs",
                               "Print QR codes", "Manage attendance"},
    "On Mismanagement": {"Post event chat message"},
    "GM (Grand Master)": set(ALL_FN) - {"Assign app-access flags"},
    "VGM (Vice GM)": set(ALL_FN) - {"Assign app-access flags"},
    "RA (Religious Advisor)": {"View payment report", "Take payment (check-in)", "Bulk payment",
                               "Manage receipts (expenses)", "Manage attendance",
                               "Award list (drinks)", "Manage Down Downs", "Post event chat message"},
    "Beer Meister": {"Manage receipts (expenses)", "Award list (drinks)", "Manage Down Downs"},
    "Hash Flash": {"Review / approve photos", "Edit photo status / caption", "Batch / view all photos"},
    "On Sec": {"Create / edit runs", "Manage attendance", "Copy RSVPs between runs",
               "Manage members (roster)", "View invite codes", "Post event chat message"},
    "Song Meister": {"Manage songs"},
    "Trail Master": {"Create / edit runs", "Print QR codes", "Manage attendance",
                     "Copy RSVPs between runs", "PackTrack trim (admin)"},
    "Hare Raiser": {"Create / edit runs", "Print QR codes"},
    "Hash Cash": {"View payment report", "Take payment (check-in)", "Bulk payment",
                  "Manage receipts (expenses)"},
    "Scribe": {"Write / save Hash Trash", "View Hash Trash drafts"},
    "Web Meister": {"Batch / view all photos", "Write / save Hash Trash", "View Hash Trash drafts"},
    "Hash Hugs": {"Post event chat message"},
    "Hash Ho": set(),          # unsure of this role — left blank for James
    "Haberdasher": {"Manage receipts (expenses)"},
    "Hash Sweep": set(),       # trail role, not admin
    "Hash Trash": {"Write / save Hash Trash", "View Hash Trash drafts"},
    "Hash Bank": {"View payment report", "Take payment (check-in)", "Bulk payment",
                  "Manage receipts (expenses)"},
    "Event Meister": {"Create / edit runs", "Print QR codes", "Manage attendance",
                      "Copy RSVPs between runs"},
    "Communications": {"Post event chat message", "Write / save Hash Trash", "View Hash Trash drafts"},
    "Other": set(),
}
YES_FILL = PatternFill("solid", fgColor="C6EFCE")
YES_FONT = Font(bold=True, color="006100", size=10)

# ---- styling helpers ------------------------------------------------------
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILLS = ["4472C4", "2E75B6", "1F6F54", "7A5195", "BC5090", "996515", "3C6E71", "5B5F97", "8C4A3B"]
ROLE_FONT = Font(bold=True, size=10)
REF_FILL = PatternFill("solid", fgColor="FFE699")   # amber for the admin reference rows
ROLE_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center")
VHEAD = Alignment(horizontal="center", vertical="bottom", text_rotation=90, wrap_text=True)

wb = openpyxl.Workbook()

# ===========================================================================
# Sheet 1 — Assign Privileges (the matrix to fill in)
# ===========================================================================
ws = wb.active
ws.title = "Assign Privileges"

n_cols = 1 + len(functions)
last_col = get_column_letter(n_cols)

# Row 1 — instruction banner
ws.merge_cells(f"A1:{last_col}1")
c = ws["A1"]
c.value = ('PERMISSIONS MATRIX — PRE-FILLED WITH BEST-GUESS DEFAULTS (green "Y" = has access); review & correct. '
           'Amber rows: Super Admin / Admin are AppAccessFlags; "Hare (this run only) *" is RUN-SCOPED — a designated '
           'hare gets those functions only for the run they are haring, not kennel-wide. '
           '"Hash Ho" left blank (unsure of the role). These are ROLE defaults; the per-hasher flag override is separate. '
           'See the "Function Reference" tab for what each function is and how it is gated TODAY.')
c.font = Font(bold=True, size=11, color="9C0006")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 46

# Row 2 — group headers (merged per group); Row 3 — function names
col = 2
gi = 0
group_spans = {}
for gname, items in groups:
    start = col
    for it in items:
        fc = ws.cell(row=3, column=col, value=it[0])
        fc.font = Font(bold=True, size=9)
        fc.alignment = VHEAD
        fc.border = BORDER
        col += 1
    end = col - 1
    ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    gc = ws.cell(row=2, column=start, value=gname)
    gc.font = HDR_FONT
    gc.alignment = CENTER
    gc.fill = PatternFill("solid", fgColor=GROUP_FILLS[gi % len(GROUP_FILLS)])
    gc.border = BORDER
    # tint the function header cells to match the group
    for cc in range(start, end + 1):
        ws.cell(row=3, column=cc).fill = PatternFill("solid", fgColor=GROUP_FILLS[gi % len(GROUP_FILLS)])
        ws.cell(row=3, column=cc).font = Font(bold=True, size=9, color="FFFFFF")
    gi += 1

corner = ws.cell(row=3, column=1, value="Mismanagement Role  ↓   /   Function  →")
corner.font = Font(bold=True, size=9)
corner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
corner.border = BORDER
ws.cell(row=2, column=1).border = BORDER
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 150

# Data rows — reference (admin) then mm roles
row = 4
def write_role_row(name, fill):
    global row
    rc = ws.cell(row=row, column=1, value=name)
    rc.font = ROLE_FONT
    rc.fill = fill
    rc.alignment = Alignment(horizontal="left", vertical="center")
    rc.border = BORDER
    granted = FILL.get(name, set())
    for cidx in range(2, n_cols + 1):
        fn_short = functions[cidx - 2][1]
        is_yes = fn_short in granted
        cell = ws.cell(row=row, column=cidx, value="Y" if is_yes else "")
        cell.alignment = CENTER
        cell.border = BORDER
        if is_yes:
            cell.fill = YES_FILL
            cell.font = YES_FONT
        elif fill is REF_FILL:
            cell.fill = REF_FILL
    ws.row_dimensions[row].height = 18
    row += 1

for r in appaccess_ref:
    write_role_row(r, REF_FILL)
for r in mm_roles:
    write_role_row(r, ROLE_FILL)

# Data validation dropdown Y on the matrix body
dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv.add(f"B4:{last_col}{row-1}")
ws.add_data_validation(dv)

# widths + freeze
ws.column_dimensions["A"].width = 30
for cidx in range(2, n_cols + 1):
    ws.column_dimensions[get_column_letter(cidx)].width = 4.5
ws.freeze_panes = "B4"

# ===========================================================================
# Sheet 2 — Function Reference
# ===========================================================================
ws2 = wb.create_sheet("Function Reference")
headers = ["Feature area", "Function", "Stored procedure(s)", "Current SERVER gate", "Current APP (UI) gate", "Notes"]
for i, h in enumerate(headers, start=1):
    cc = ws2.cell(row=1, column=i, value=h)
    cc.font = HDR_FONT
    cc.fill = PatternFill("solid", fgColor="404040")
    cc.alignment = Alignment(horizontal="left", vertical="center")
    cc.border = BORDER
r = 2
gi = 0
for gname, items in groups:
    for it in items:
        vals = [gname, it[0], it[1], it[2], it[3], it[4]]
        for i, v in enumerate(vals, start=1):
            cc = ws2.cell(row=r, column=i, value=v)
            cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cc.border = BORDER
            cc.font = Font(size=10)
            if i == 1:
                cc.fill = PatternFill("solid", fgColor=GROUP_FILLS[gi % len(GROUP_FILLS)])
                cc.font = Font(size=10, bold=True, color="FFFFFF")
            if "UNGATED" in str(v) or "CRITICAL" in str(v) or "P0" in str(v):
                cc.font = Font(size=10, bold=True, color="9C0006")
        r += 1
    gi += 1
widths2 = [16, 30, 34, 28, 26, 22]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# footnote about admin* on sheet 2
ws2.cell(row=r + 1, column=1, value="* 'admin' = SuperAdmin | Admin, but the current mask (0x40000081) also wrongly includes ManagePublicWebContent (0x80) — see audit.").font = Font(italic=True, size=9, color="7F6000")

# NOTE: this regenerates the BLANK template. Once docs/permissions_matrix.xlsx has
# been hand-filled with role assignments, do NOT re-run this against it — it will
# wipe the marks. Regenerate to a temp path and merge new functions/roles by hand,
# or update the hand-edited workbook directly.
import os
out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "docs", "permissions_matrix.xlsx")
wb.save(out)
print("saved", out)
print("roles:", len(appaccess_ref) + len(mm_roles), "functions:", len(functions))
