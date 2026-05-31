"""
Generates the Harrier Central 3.0 test plan spreadsheet.
Run: python3 tools/generate_test_plan.py
Output: tools/HC3_Test_Plan.xlsx
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_BG   = "1E3A5F"   # dark navy
HEADER_FG   = "FFFFFF"
CAT_COLORS  = {           # category row highlight (light tints)
    "BOOT":               "E8F4FD",
    "AUTH":               "FEF3E2",
    "MIGRATION":          "E8F8F5",
    "CHAT":               "F3E8FF",
    "PHOTOS":             "FFF0E8",
    "PACKTRACK":          "E8FFE8",
    "SYNC":               "E8F4FD",
    "RUNS":               "FFFBE8",
    "ADMIN_RUN":          "F8E8FF",
    "ADMIN_KENNEL":       "E8F0FF",
    "PERMISSIONS":        "FFE8E8",
    "PORTAL_RUN":         "F0FFF0",
    "PORTAL_KENNEL":      "FFF8E8",
    "PORTAL_CHAT":        "F8E8F8",
    "PORTAL_PHOTOS":      "FFFEE8",
    "PORTAL_PERMISSIONS": "E8FFFF",
    "NOTIFICATIONS":      "F8F8E8",
    "NEWSFLASH":          "FFE8F4",
    "SECURITY":           "FFE8E8",
    "PROFILE":            "E8F8FF",
    "REGRESSION":         "F4F4F4",
}
ALT_ROW     = "F9FBFD"   # alternating row tint
BORDER_COL  = "BDD7EE"

# ── Test data ─────────────────────────────────────────────────────────────────
# Columns: (serial, name, category, platform, description)
# Platform: APP | PORTAL | APP+PORTAL
TESTS = [

    # ── BOOT ──────────────────────────────────────────────────────────────────
    ("TC-001", "Cold launch — good network",           "BOOT", "APP",
     "Launch the app from a cold state on a good Wi-Fi connection. Verify the app reaches main content within 5 seconds and the loading indicator does not stick."),
    ("TC-002", "Cold launch — no network",             "BOOT", "APP",
     "Launch the app with airplane mode enabled. Verify the offline-mode ribbon appears and the app loads from the local SQLite database without crashing."),
    ("TC-003", "Cold launch — slow network",           "BOOT", "APP",
     "Launch on a throttled connection (3G or equivalent). Verify a 30-second GPS wait cap is respected; the app proceeds to content rather than hanging indefinitely."),
    ("TC-004", "Boot logger: no credentials in logs",  "BOOT", "APP",
     "After a normal boot, inspect the debug log (Xcode console / adb logcat). Confirm that no accessToken, deviceId, or deviceSecret strings appear in any log entry."),
    ("TC-005", "Boot instrumentation timing",          "BOOT", "APP",
     "In a debug build, confirm [BOOT] timing markers appear in sequence from app start through to app content visible. Identify any step taking > 2 seconds."),
    ("TC-006", "Return from background — short absence", "BOOT", "APP",
     "Minimise the app for 30 seconds then foreground it. Verify no reload occurs and the user stays on the same screen."),
    ("TC-007", "Return from background — long absence", "BOOT", "APP",
     "Minimise the app overnight (8+ hours) then foreground it. Verify a delta sync runs silently and the run list refreshes without a full reload screen."),
    ("TC-008", "First launch after install",           "BOOT", "APP",
     "Fresh install. Verify the intro slider displays, no warm-reload dialog appears, and the user reaches the QR pairing screen."),

    # ── AUTH ──────────────────────────────────────────────────────────────────
    ("TC-009", "QR code pairing — new device",         "AUTH", "APP+PORTAL",
     "On a new install, scan the QR code from the portal My Profile page. Verify the device is authorised, a deviceSecret is generated, and the user lands on the main navigation page."),
    ("TC-010", "QR code pairing — replace existing device", "AUTH", "APP+PORTAL",
     "On a device already registered to another user, scan a different user's QR code. Verify the old account is replaced and the new user's profile loads correctly."),
    ("TC-011", "Reset code re-auth",                   "AUTH", "APP",
     "Simulate a warm-reload by forcing DB_VERSION mismatch. Verify the app calls hcapp_authorizeDevice with the stored reset code, receives new credentials, and boots normally."),
    ("TC-012", "Device not registered error",          "AUTH", "APP",
     "Attempt to boot after the device record has been deleted server-side. Verify a 'Device Not Authorised' dialog appears prompting QR re-scan, and the app does not crash."),
    ("TC-013", "Token expiry — mid-session",           "AUTH", "APP",
     "Observe normal usage; tokens regenerate on each call. Verify no 'Invalid Access Token' errors appear during normal app use or after returning from background."),
    ("TC-014", "iOS background suspension token fix",  "AUTH", "APP",
     "During a live run on iOS, lock the screen for 10+ minutes, then unlock and return to the app. Verify GPS positions continue sending and no token errors appear in the run log."),
    ("TC-015", "Portal login — QR scan flow",          "AUTH", "PORTAL",
     "Open the portal login page on a new browser session. Scan the displayed QR code with the authorised mobile app (portal confirmation). Verify login completes and the portal main page loads."),
    ("TC-016", "Portal session persistence",           "AUTH", "PORTAL",
     "Log into the portal, close the browser tab, and reopen the portal URL. Verify the session is restored without requiring re-login (within session timeout period)."),

    # ── MIGRATION ─────────────────────────────────────────────────────────────
    ("TC-017", "Warm reload — happy path",             "MIGRATION", "APP",
     "Install 2.6.8+1140 over any prior version. Launch. Verify: (1) warm-reload dialog shows 'Account Load Successful', (2) user profile loads completely after reload, (3) all run data is present."),
    ("TC-018", "Warm reload — no data loss",           "MIGRATION", "APP",
     "After the warm reload triggered by TC-017, open each major section: run list, kennel admin, chat, profile. Verify all data is present and no sections show empty or error states."),
    ("TC-019", "Warm reload — no second trigger",      "MIGRATION", "APP",
     "After the warm reload in TC-017, force-close and relaunch the app. Verify no second warm-reload dialog appears and the app boots normally."),
    ("TC-020", "Storage type in Imprint — encrypted",  "MIGRATION", "APP",
     "After successful migration (TC-017), open drawer → Imprint. Verify the Software version block shows 'Storage: encrypted'."),
    ("TC-021", "Storage type in Imprint — legacy",     "MIGRATION", "APP",
     "On a fresh install (no prior version), open Imprint before any warm-reload. Verify the Software version block shows 'Storage: legacy'."),
    ("TC-022", "Dialog title — encrypted path",        "MIGRATION", "APP",
     "On the warm-reload dialog after a successful secure-storage migration, verify the title reads 'Account Load Successful' (not 'Profile Load Successful')."),
    ("TC-023", "Reset code survives migration",        "MIGRATION", "APP",
     "After the warm reload, force another warm reload (if testable). Verify the app can still re-authorise using the reset code stored in secure storage without requesting QR re-scan."),

    # ── CHAT ──────────────────────────────────────────────────────────────────
    ("TC-024", "Send message — app to app",            "CHAT", "APP",
     "User A sends a chat message on an event. Verify User B (different device, same event) receives the message within 5 seconds via FCM push."),
    ("TC-025", "Send message — portal to app",         "CHAT", "APP+PORTAL",
     "Send a message from the portal chat panel. Verify the message appears on the app within 5 seconds for all subscribed users."),
    ("TC-026", "Send message — app to portal",         "CHAT", "APP+PORTAL",
     "Send a message from the app. Verify the message appears in the portal chat panel within 5 seconds."),
    ("TC-027", "Message status: Sending → Sent",       "CHAT", "APP",
     "Send a message on a good network. Verify the message shows 'Sending' immediately, then transitions to 'Sent' (single tick) after the API call confirms success."),
    ("TC-028", "Message status: Sent → Delivered",     "CHAT", "APP",
     "Send a message to a recipient who is online. Verify the message status transitions from 'Sent' to 'Delivered' (double tick) after the recipient's FCM echo is received."),
    ("TC-029", "Message status: Sending → Error",      "CHAT", "APP",
     "Send a message with no network connection. Verify the message shows an error state and does not remain stuck in 'Sending'."),
    ("TC-030", "Delta-fetch on app resume",            "CHAT", "APP",
     "With app backgrounded, send 3 messages from another device. Bring the app back to foreground. Verify all 3 messages are fetched within ~5 seconds of resuming."),
    ("TC-031", "No duplicate messages",                "CHAT", "APP",
     "Send a burst of 5 messages rapidly. Verify each message appears exactly once in the chat list; no duplicates introduced by the pending fetch queue."),
    ("TC-032", "Chat unread count badge",              "CHAT", "APP",
     "Receive messages in a chat while on a different page. Verify the chat badge count increments correctly and clears when the chat page is opened."),
    ("TC-033", "ChatStripWidget — live run embed",     "CHAT", "APP",
     "During a live run, verify the ChatStripWidget shows the latest messages in read-only mode on the Live Run Tools page. New messages appear without navigating to full chat."),
    ("TC-034", "Chat — FCM not registered device",     "CHAT", "APP",
     "Use a device that has not registered for FCM (or has revoked permission). Send a message to that device. Verify graceful handling with no crash or infinite retry loop."),
    ("TC-035", "Portal chat — delivery receipt dialog","CHAT", "PORTAL",
     "From the portal chat panel, send a message. Verify the send button shows recipient count confirmation. Verify double-tick appears once recipients' FCM echo is received."),
    ("TC-036", "Portal chat — missed app messages",    "CHAT", "APP+PORTAL",
     "Send messages from the app while the portal chat is open. Verify the portal's delta-fetch picks up app-sent messages within 30 seconds without requiring a page reload."),

    # ── PHOTOS ────────────────────────────────────────────────────────────────
    ("TC-037", "Take photo during live run",           "PHOTOS", "APP",
     "During an active live run, tap the Take Photo button. Verify the camera opens, a photo can be captured, and the 4-step review flow (review → crop → share/upload) appears."),
    ("TC-038", "Photo upload — success",               "PHOTOS", "APP",
     "Complete the photo review flow. Verify the photo uploads to Azure Blob Storage at path <kennel-slug>/<run-id>/<filename>.jpg and a success indicator is shown."),
    ("TC-039", "Photo upload — SAS domain validation", "PHOTOS", "APP",
     "Verify the SAS token URL used for upload points to harriercentral.blob.core.windows.net. Any URL with a different domain should show an error, not attempt an upload."),
    ("TC-040", "Photo review — discard",               "PHOTOS", "APP",
     "In the 4-step review flow, tap Discard. Verify the photo is not uploaded, no blob is created, and the camera is dismissed cleanly."),
    ("TC-041", "Photo review — crop",                  "PHOTOS", "APP",
     "In the review flow, tap Edit/Crop. Verify the crop tool opens, a crop can be applied, and the cropped image is used for the upload."),
    ("TC-042", "Save to camera roll — ON",             "PHOTOS", "APP",
     "Enable 'Save photos to camera roll' in My Profile. Take a photo during a run. Verify the photo appears in the device camera roll in addition to being uploaded."),
    ("TC-043", "Save to camera roll — OFF",            "PHOTOS", "APP",
     "Disable 'Save photos to camera roll' in My Profile. Take a photo during a run. Verify the photo does NOT appear in the device camera roll."),
    ("TC-044", "Photo caption",                        "PHOTOS", "APP",
     "During the photo review flow, add a caption. Verify the caption is stored with the photo and visible in the photo viewer."),
    ("TC-045", "PHO marker appears on run map",        "PHOTOS", "APP",
     "After uploading photos during a run, open the run map. Verify camera-frame thumbnail markers (PHO) appear at the correct GPS coordinates where photos were taken."),
    ("TC-046", "PHO marker — zoom scaling",            "PHOTOS", "APP",
     "On the run map with photo markers, zoom in and out. Verify markers scale smoothly: ~25px at full-run zoom, ~360px at max zoom."),
    ("TC-047", "PHO marker — loading indicator",       "PHOTOS", "APP",
     "When photo markers are loading thumbnails, verify a count-up loading indicator is shown inside the camera frame until the thumbnail resolves."),
    ("TC-048", "PHO marker — spiderfying",             "PHOTOS", "APP",
     "Take two photos at the same GPS location. Open the map. Verify the co-located markers spiderfy (fan out) when tapped rather than stacking invisibly."),
    ("TC-049", "PHO marker — tap to view",             "PHOTOS", "APP",
     "Tap a photo marker on the run map. Verify the full-resolution photo opens in a viewer overlay."),
    ("TC-050", "Portal photo review — load pending",   "PHOTOS", "APP+PORTAL",
     "After a photo is uploaded from the app, open the portal → navigate to the run → Review Photos. Verify the photo appears in the Pending tab."),
    ("TC-051", "Portal photo review — Approve (Share)","PHOTOS", "APP+PORTAL",
     "In the portal photo review screen, select a pending photo and tap 'Approve (Share)'. Verify the photo moves to the Reviewed tab with status 'Shared' and is now visible to other users in the app."),
    ("TC-052", "Portal photo review — Delete",         "PHOTOS", "APP+PORTAL",
     "In portal photo review, tap Delete on a photo. Verify the photo is soft-deleted and no longer appears in the app's run photo list."),
    ("TC-053", "Portal photo review — Keep Private",   "PHOTOS", "APP+PORTAL",
     "In portal photo review, tap Keep Private. Verify status changes to Private and the photo is only visible to the uploader in the app."),
    ("TC-054", "Portal photo review — Gallery",        "PHOTOS", "APP+PORTAL",
     "In portal photo review, set a photo to Gallery. Verify status updates and the photo appears in the kennel gallery view in the app."),
    ("TC-055", "Portal photo review — Home Gallery",   "PHOTOS", "APP+PORTAL",
     "In portal photo review, set a photo to Home Gallery. Verify status updates and the photo appears prominently on the kennel home view."),
    ("TC-056", "Portal photo review — Event Cover",    "PHOTOS", "APP+PORTAL",
     "In portal photo review, set a photo to Event Cover. Verify HC.Event.EventCoverPhotoUrl is updated and the photo appears as the run's cover image in the app."),
    ("TC-057", "Photo review pagination (Next/Prev)",  "PHOTOS", "PORTAL",
     "When multiple pending photos exist for a run, use the Next and Previous navigation in the portal review screen. Verify photos cycle correctly and current index is shown ('2 of 5' etc)."),
    ("TC-058", "Portal photo review — auth enforcement","PHOTOS", "APP+PORTAL",
     "Log in as a portal user who does NOT have authIsAdmin or authCanManagePublicWebContent for the kennel. Verify the Review Photos button is not shown or returns an auth error."),

    # ── PACKTRACK ─────────────────────────────────────────────────────────────
    ("TC-059", "Start live run",                       "PACKTRACK", "APP",
     "From the run detail page for an active event, tap Start Live Run. Verify GPS starts recording and the Live Run shell activates with the Tools page accessible."),
    ("TC-060", "Pause tracking",                       "PACKTRACK", "APP",
     "During an active live run, tap Pause. Verify GPS broadcasting stops, the pause button changes to Resume, and no positions are sent to the server while paused."),
    ("TC-061", "Resume after manual pause",            "PACKTRACK", "APP",
     "After pausing (TC-060), tap Resume. Verify GPS broadcasting resumes immediately and the track continues from the paused position."),
    ("TC-062", "Auto-pause: stationary > threshold",   "PACKTRACK", "APP",
     "Start a live run and remain stationary. Verify auto-pause does NOT trigger (auto-pause is proximity-based on resumption, not triggered by being stationary)."),
    ("TC-063", "Auto-resume: move 100m+ while paused", "PACKTRACK", "APP",
     "Pause a live run, then walk more than 100 metres away. Verify tracking auto-resumes and a visual indicator confirms the auto-resume event."),
    ("TC-064", "Auto-resume: stay within 100m while paused","PACKTRACK","APP",
     "Pause a live run and move only 80 metres. Verify tracking does NOT auto-resume."),
    ("TC-065", "End Run — confirmation dialog",        "PACKTRACK", "APP",
     "During a live run, tap End Run. Verify a confirmation dialog appears asking the user to confirm before ending."),
    ("TC-066", "End Run — confirm",                    "PACKTRACK", "APP",
     "Confirm the End Run dialog. Verify tracking stops, the live run shell exits, and the run data is saved to the server."),
    ("TC-067", "End Run — cancel",                     "PACKTRACK", "APP",
     "Dismiss the End Run dialog without confirming. Verify the live run continues uninterrupted."),
    ("TC-068", "Background GPS — iOS",                 "PACKTRACK", "APP",
     "During a live run on iOS, lock the device screen. After 5 minutes, unlock and return to the app. Verify GPS positions were sent during the background period (check position count on server)."),
    ("TC-069", "Background GPS — Android",             "PACKTRACK", "APP",
     "During a live run on Android, minimise the app. After 5 minutes, return to the app. Verify GPS positions were sent during the background period."),
    ("TC-070", "Other runners on map",                 "PACKTRACK", "APP",
     "With two devices on the same live run, verify each device can see the other's position marker updating in near-real-time on the run map."),
    ("TC-071", "GPS track filter — preserve typed points","PACKTRACK","APP",
     "Start a live run on a complex route with direction changes. Verify that typed track points (e.g. checkpoints) are not filtered out by the GPS smoothing filter."),
    ("TC-072", "Distance preference — km",             "PACKTRACK", "APP",
     "Set distance preference to km. Verify all distances in the live run UI (distance covered, split times) show in km."),
    ("TC-073", "Distance preference — miles",          "PACKTRACK", "APP",
     "Set distance preference to miles. Verify all distances in the live run UI show in miles. Verify the preference auto-saves without requiring a separate Save action."),
    ("TC-074", "Live run nav bar — selected icon colour","PACKTRACK","APP",
     "During a live run, verify the selected navigation bar icon is white (not the default tint colour)."),

    # ── SYNC ──────────────────────────────────────────────────────────────────
    ("TC-075", "Common domain — initial sync",         "SYNC", "APP",
     "On first login, verify the common domain sync completes: runs, hashers, kennels, songs, HKM, HEM and payment rows are all populated in the local SQLite DB."),
    ("TC-076", "Common domain — delta sync",           "SYNC", "APP",
     "After initial sync, a new run is created in the portal. Trigger a sync from the app. Verify the new run appears in the run list without a full re-sync."),
    ("TC-077", "Kennel follow — force replicate",      "SYNC", "APP",
     "Follow a kennel the user has never been a member of. Verify a force-replicate syncs all historical events for that kennel so the full run history is available."),
    ("TC-078", "Kennel admin entry — fresh sync",      "SYNC", "APP",
     "Enter kennel admin for the first time. Verify the kennel domain tables are wiped and a full kennel sync runs before admin screens are accessible."),
    ("TC-079", "Event admin entry — fresh sync",       "SYNC", "APP",
     "Enter run admin (event domain) for a run. Verify event tables are wiped and a full event sync runs before the admin screen shows data."),
    ("TC-080", "Sync debounce — rapid navigation",     "SYNC", "APP",
     "Navigate quickly between several runs in the list. Verify sync does not fire for every tap; debouncing prevents redundant concurrent syncs."),
    ("TC-081", "Sync hang guard",                      "SYNC", "APP",
     "Simulate a slow or non-responding server during sync. Verify the sync times out gracefully (does not hang indefinitely) and the app shows an appropriate error or offline state."),

    # ── RUNS ──────────────────────────────────────────────────────────────────
    ("TC-082", "Run list — upcoming runs",             "RUNS", "APP",
     "Open the run list. Verify upcoming runs appear in chronological order with correct event names, dates, and kennel logos."),
    ("TC-083", "Run list — past runs",                 "RUNS", "APP",
     "Scroll to past runs. Verify recent past runs appear, attendance state is shown, and payment status is indicated where applicable."),
    ("TC-084", "Run list — refreshes after edit",      "RUNS", "APP",
     "Edit a run in the portal (change name or date). Return to the app run list. Verify the updated run reflects the changes without a manual restart."),
    ("TC-085", "Run detail — all sections load",       "RUNS", "APP",
     "Open a past run detail page. Verify all sections load: attendance, map/track, photos, chat, and admin buttons (if admin)."),
    ("TC-086", "Run detail — stale controller fix",    "RUNS", "APP",
     "Navigate to run A detail, return to list, immediately navigate to run B detail. Verify run B's data loads correctly and no data from run A is shown (stale controller prevented)."),
    ("TC-087", "Run RSVP",                             "RUNS", "APP",
     "RSVP for a future run. Verify the RSVP state persists after app restart. Verify the RSVP count on the run detail page updates."),
    ("TC-088", "Hash credit display on check-in",      "RUNS", "APP",
     "For a user with available hash credit, open a run check-in sheet. Verify the credit amount is displayed on their pack list item and the card height adjusts correctly."),
    ("TC-089", "Kennel website QR",                    "RUNS", "APP",
     "From a kennel's page in the app, access the kennel website QR. Scan the QR with a camera. Verify it navigates to the correct kennel website URL (www.hashruns.org/<slug> or custom domain)."),
    ("TC-090", "Promoted event display",               "RUNS", "APP",
     "In the portal, mark a run as a promoted event (IsPromotedEvent checkbox). Verify it appears with appropriate promotion styling in the app run list."),

    # ── ADMIN_RUN ─────────────────────────────────────────────────────────────
    ("TC-091", "Mark attendance — self",               "ADMIN_RUN", "APP",
     "As a regular user, mark yourself as attending a run. Verify attendance state updates in the app and syncs to the server."),
    ("TC-092", "Mark attendance — another user (admin)","ADMIN_RUN","APP",
     "As a run admin, mark another user as attending. Verify the attendance record is saved and the user's HEM row updates."),
    ("TC-093", "Bulk attendance",                      "ADMIN_RUN", "APP",
     "As a run admin, use bulk attendance to mark multiple hashers at once. Verify all records are saved correctly in a single round-trip."),
    ("TC-094", "Process payment",                      "ADMIN_RUN", "APP",
     "As a hash cash admin, process a payment for an attendee. Verify the payment record is saved, the attendee's balance updates, and the payment appears in the payment report."),
    ("TC-095", "Payment report",                       "ADMIN_RUN", "APP",
     "As a hash cash admin, open the payment report for a run. Verify all attendees are listed with correct payment status, amounts, and credit balances."),
    ("TC-096", "RSVP copy between runs",               "ADMIN_RUN", "APP",
     "As a run admin, copy RSVPs from a previous run to an upcoming run. Verify RSVPs are correctly copied and only for hashers in the correct kennel."),
    ("TC-097", "Run admin button layout",              "ADMIN_RUN", "APP",
     "Open the run admin page. Verify the button tiles are displayed at 110×110 with spaceAround layout (not stretching or overflowing)."),

    # ── ADMIN_KENNEL ──────────────────────────────────────────────────────────
    ("TC-098", "Kennel admin — member list",           "ADMIN_KENNEL", "APP",
     "Enter kennel admin. Verify the full member list loads including names, run counts, and payment status."),
    ("TC-099", "Kennel admin — admin for unfollowed kennel","ADMIN_KENNEL","APP",
     "Attempt to enter kennel admin for a kennel the user is admin of but does not currently follow. Verify the app auto-follows the kennel and force-replicates event history before showing admin screens."),
    ("TC-100", "Kennel status — Active",               "ADMIN_KENNEL", "APP+PORTAL",
     "Set kennel status to Active in the portal. Verify the kennel appears normally in the app kennel list and discovery."),
    ("TC-101", "Kennel status — Inactive-Visible",     "ADMIN_KENNEL", "APP+PORTAL",
     "Set kennel status to Inactive-Visible in the portal. Verify the kennel is visible in the app but marked as inactive."),
    ("TC-102", "Kennel status — Inactive-Hidden",      "ADMIN_KENNEL", "APP+PORTAL",
     "Set kennel status to Inactive-Hidden in the portal. Verify the kennel does not appear in discovery or search but is accessible via direct link."),
    ("TC-103", "Kennel status — Defunct",              "ADMIN_KENNEL", "APP+PORTAL",
     "Set kennel status to Defunct in the portal. Verify appropriate handling in the app (hidden from active listings, historical data preserved)."),
    ("TC-104", "Mismanagement roles save",             "ADMIN_KENNEL", "APP",
     "In kennel admin, update a member's mismanagement roles. Save. Re-open the member's admin panel. Verify the roles persist correctly across the save/reload cycle."),

    # ── PERMISSIONS ───────────────────────────────────────────────────────────
    ("TC-105", "Grant authCanManageSongs",              "PERMISSIONS", "APP+PORTAL",
     "In the portal, grant authCanManageSongs (0x0100) to a user. In the app, verify that user can now access the Manage Songs screen and edit songs for the kennel."),
    ("TC-106", "Revoke authCanManageSongs",             "PERMISSIONS", "APP+PORTAL",
     "In the portal, revoke authCanManageSongs from a user. Verify the Manage Songs option disappears from the app for that user after the next sync."),
    ("TC-107", "Grant authCanManagePublicWebContent",   "PERMISSIONS", "APP+PORTAL",
     "In the portal, grant authCanManagePublicWebContent (0x0080) to a user. Verify the user gains access to web content management features in both the app and the portal."),
    ("TC-108", "AppAccess page — Manage Songs toggle visible","PERMISSIONS","APP",
     "Open the AppAccess page in the app for a kennel where the current user is admin. Verify the Manage Songs toggle is present and functional."),
    ("TC-109", "AppAccess page — flags save correctly","PERMISSIONS","APP",
     "Toggle multiple AppAccess flags. Save. Re-open the AppAccess page. Verify all toggled flags are persisted correctly."),
    ("TC-110", "Portal permission wiring — admin only","PERMISSIONS","PORTAL",
     "Log into the portal as a non-admin user. Verify that admin-only sections (member management, platform admin) are not accessible."),

    # ── PORTAL_RUN ────────────────────────────────────────────────────────────
    ("TC-111", "Create new run",                       "PORTAL_RUN", "APP+PORTAL",
     "In the portal run editor, create a new event with all required fields (name, date, location). Save. Verify the run appears in the app run list after next sync."),
    ("TC-112", "Edit run — name and date",             "PORTAL_RUN", "APP+PORTAL",
     "In the portal, edit an existing run's name and date. Save. Verify the updated details appear in the app."),
    ("TC-113", "Edit run — location lookup",           "PORTAL_RUN", "PORTAL",
     "In the portal run editor, use the location lookup (auto-search). Verify results appear, smart matching ranks correct results, and selecting a result fills the location fields."),
    ("TC-114", "Edit run — geocoding fallback",        "PORTAL_RUN", "PORTAL",
     "Use the auto-locate button in the portal run editor for a location in an obscure area. Verify geocoding falls back to kennel country when a narrow region search returns no results."),
    ("TC-115", "Edit run — IsPromotedEvent",           "PORTAL_RUN", "PORTAL",
     "Toggle the IsPromotedEvent checkbox in the portal run editor. Save. Verify the flag is persisted and reflects correctly on the run detail in the portal."),
    ("TC-116", "Edit run — EventGeographicScope",      "PORTAL_RUN", "PORTAL",
     "Set EventGeographicScope in the portal run editor. Save. Verify the value persists and is retrievable via the run detail endpoint."),
    ("TC-117", "Delete run",                           "PORTAL_RUN", "APP+PORTAL",
     "Soft-delete a run in the portal. Verify the run no longer appears in the app run list after next sync."),

    # ── PORTAL_KENNEL ─────────────────────────────────────────────────────────
    ("TC-118", "Edit kennel — basic details",          "PORTAL_KENNEL", "APP+PORTAL",
     "In the portal kennel editor, update the kennel name and short name. Save. Verify the changes appear in the app after sync."),
    ("TC-119", "Edit kennel — requires admin flag",    "PORTAL_KENNEL", "PORTAL",
     "Log in as a portal user without the canManageKennel flag. Attempt to edit kennel settings. Verify access is denied with an appropriate error."),
    ("TC-120", "Platform Admin tab — CanEditKennel",   "PORTAL_KENNEL", "PORTAL",
     "Log in as a platform admin (CanEditKennel privilege). Verify the Platform Admin tab is visible in the kennel editor and allows editing platform-level settings."),
    ("TC-121", "Platform Admin tab — non-platform-admin","PORTAL_KENNEL","PORTAL",
     "Log in as a regular kennel admin (not a platform admin). Verify the Platform Admin tab is NOT visible in the kennel editor."),
    ("TC-122", "Kennel editor — check-in sheet timeout","PORTAL_KENNEL","PORTAL",
     "Open the kennel check-in sheet in the portal with a kennel that has a slow or broken logo URL. Verify the sheet loads within a reasonable time (logo fetch has a timeout) and does not hang."),

    # ── PORTAL_CHAT ───────────────────────────────────────────────────────────
    ("TC-123", "Portal send message — appears in app", "PORTAL_CHAT", "APP+PORTAL",
     "Send a chat message from the portal. Verify the message arrives in the app chat within 5 seconds via FCM."),
    ("TC-124", "Portal FCM subscription",              "PORTAL_CHAT", "PORTAL",
     "Open the portal for a kennel/event with an active chat. Verify FCM subscription is established and new messages received via push without polling."),
    ("TC-125", "Portal auto-resubscribe",              "PORTAL_CHAT", "PORTAL",
     "Let the portal session run for 30+ minutes with the chat open. Verify FCM subscription is refreshed automatically and messages continue arriving without a page reload."),
    ("TC-126", "Portal chat — Safari polling fallback","PORTAL_CHAT", "PORTAL",
     "Open the portal chat in Safari. Verify that if FCM is not supported, a 30-second polling fallback kicks in and messages arrive (possibly with slight delay)."),

    # ── PORTAL_PHOTOS ─────────────────────────────────────────────────────────
    ("TC-127", "Portal photo review — auth via publicKennelId","PORTAL_PHOTOS","APP+PORTAL",
     "Open the portal photo review page for any kennel. Verify it loads without an 'unauthorised' error. (Regression for the publicKennelId fix — previously failed due to internal vs public ID mismatch.)"),
    ("TC-128", "Portal photo review — batch status update","PORTAL_PHOTOS","PORTAL",
     "Apply a status action (e.g. Approve) to a photo. Verify the batchUpdatePhotoStatus SP is called with the correct publicKennelId and the status updates in the database."),
    ("TC-129", "Portal photo review — Reviewed tab",   "PORTAL_PHOTOS", "PORTAL",
     "After reviewing several photos, switch to the Reviewed tab. Verify previously actioned photos appear with their current status (Shared, Private, Gallery, etc)."),

    # ── PORTAL_PERMISSIONS ────────────────────────────────────────────────────
    ("TC-130", "Portal member list — load all members","PORTAL_PERMISSIONS","PORTAL",
     "In the portal, open the kennel member list. Verify all members load with their current AppAccessFlags and mismanagement roles displayed."),
    ("TC-131", "Portal — grant permission saves correctly","PORTAL_PERMISSIONS","APP+PORTAL",
     "In the portal, grant a new permission flag to a member. Save. In the app, trigger a kennel sync. Verify the member's new permission is reflected in the app."),
    ("TC-132", "Portal — revoke permission saves correctly","PORTAL_PERMISSIONS","APP+PORTAL",
     "In the portal, revoke a permission flag from a member. Save. In the app, trigger a kennel sync. Verify the permission is removed in the app."),

    # ── NOTIFICATIONS ─────────────────────────────────────────────────────────
    ("TC-133", "Push notification — new run announcement","NOTIFICATIONS","APP",
     "Create a new run in the portal and trigger a run announcement FCM push. Verify the push notification arrives on subscribed devices, and tapping it opens the correct run detail."),
    ("TC-134", "Push notification — chat message",     "NOTIFICATIONS", "APP",
     "Send a chat message to an event while the app is backgrounded. Verify an FCM push notification arrives and tapping it navigates to the correct chat."),
    ("TC-135", "Push notification — newsflash",        "NOTIFICATIONS", "APP+PORTAL",
     "Send a newsflash from the portal. Verify an FCM push notification arrives on the app for all subscribed kennel members."),
    ("TC-136", "FCM token registration",               "NOTIFICATIONS", "APP",
     "On first launch after install, verify the device's FCM token is registered via hcapp_setFcmTokens. Verify push notifications arrive after registration."),
    ("TC-137", "FCM token refresh",                    "NOTIFICATIONS", "APP",
     "If the FCM token rotates (simulate by clearing app data), verify the app detects the new token and re-registers it with the server on next boot."),

    # ── NEWSFLASH ─────────────────────────────────────────────────────────────
    ("TC-138", "Create newsflash",                     "NEWSFLASH", "PORTAL",
     "In the portal, create a new newsflash for a kennel. Save and publish. Verify the newsflash appears in the portal newsflash list."),
    ("TC-139", "Newsflash appears in app",             "NEWSFLASH", "APP+PORTAL",
     "After publishing a newsflash in the portal (TC-138), open the app. Verify the newsflash notification appears for kennel members."),
    ("TC-140", "Respond to newsflash",                 "NEWSFLASH", "APP+PORTAL",
     "In the app, respond to a newsflash. In the portal, open the newsflash reader list. Verify the response is recorded and the read count updates."),
    ("TC-141", "Delete newsflash",                     "NEWSFLASH", "PORTAL",
     "In the portal, delete an existing newsflash. Verify it no longer appears in the portal list and is removed from the app on next sync."),
    ("TC-142", "Pending newsflashes — portal view",    "NEWSFLASH", "PORTAL",
     "Publish a newsflash and check the portal pending newsflashes view. Verify unread/pending counts are accurate and update as members respond."),

    # ── SECURITY ──────────────────────────────────────────────────────────────
    ("TC-143", "Payment URL — HTTPS enforcement",      "SECURITY", "APP",
     "Configure a kennel with an http:// payment URL (without 's'). Tap the pay button in the app. Verify no browser is opened; the URL is silently rejected."),
    ("TC-144", "Payment URL — HTTPS allowed",          "SECURITY", "APP",
     "Configure a kennel with an https:// payment URL. Tap the pay button. Verify the browser opens to the correct URL."),
    ("TC-145", "No credentials in debug logs",         "SECURITY", "APP",
     "Reproduce an API error (e.g. disconnect network mid-request). Inspect the debug log. Verify accessToken and deviceId fields are replaced with [REDACTED] in all logged request bodies."),
    ("TC-146", "IPInfo call goes through API shim",    "SECURITY", "APP",
     "On login, monitor network traffic (Charles Proxy or similar). Verify the IP geolocation request goes to harriercentralpublicapi.azurewebsites.net/api/GetIpGeoInfo, not directly to ipinfo.io."),
    ("TC-147", "UUID case normalisation — kennel match","SECURITY","APP",
     "Verify kennels are matched correctly regardless of UUID case. SQL Server returns uppercase UUIDs; the app stores them lowercase. Verify no 'kennel not found' errors occur."),
    ("TC-148", "UUID case normalisation — event match","SECURITY","APP",
     "Verify events are matched correctly regardless of UUID case. Open a run detail for a run whose UUID may have been stored in uppercase. Verify correct data loads."),
    ("TC-149", "Portal geocode — no SAS token in JS",  "SECURITY", "PORTAL",
     "Open the portal in a browser. Inspect the compiled JS bundle (search for 'sig='). Verify the Logic App SAS token is NOT present in client-side JavaScript."),
    ("TC-150", "Portal error messages — user-safe",    "SECURITY", "PORTAL",
     "Trigger a validation error in the portal (e.g. submit a run with a missing required field). Verify the displayed error message is user-friendly (errorUserMessage) and does not contain SQL constraint names, table names, or stack traces."),

    # ── PROFILE ───────────────────────────────────────────────────────────────
    ("TC-151", "Change profile photo — camera",        "PROFILE", "APP",
     "On My Profile, tap the photo and choose to take a new photo with the camera. Verify the new photo uploads, the profile header updates immediately."),
    ("TC-152", "Change profile photo — library",       "PROFILE", "APP",
     "On My Profile, tap the photo and choose from the camera library. Verify the selected photo uploads and the profile header updates."),
    ("TC-153", "Profile photo — crop",                 "PROFILE", "APP",
     "When selecting/taking a profile photo, verify the crop tool opens, a crop is applied, and the cropped image is used for the upload."),
    ("TC-154", "Email and notification preferences",   "PROFILE", "APP",
     "Change email and notification preferences on My Profile. Save. Verify preferences persist across app restart."),
    ("TC-155", "Imprint page — version info",          "PROFILE", "APP",
     "Open drawer → Imprint. Verify app name, version, build number, and database version are all displayed and accurate."),
    ("TC-156", "Imprint page — storage type",          "PROFILE", "APP",
     "Open drawer → Imprint. Verify the Storage line shows either 'Storage: encrypted' (post-migration) or 'Storage: legacy' (pre-migration or fresh install)."),

    # ── REGRESSION ────────────────────────────────────────────────────────────
    ("TC-157", "Hash history — complete run count",    "REGRESSION", "APP",
     "Open Hash History for the current user. Verify the total run count matches the expected value (cross-check against the portal or known count). Verify the list includes all attended runs across all kennels."),
    ("TC-158", "Leaderboard loads",                    "REGRESSION", "APP",
     "Open the Leaderboard. Verify it loads without error and shows current standings."),
    ("TC-159", "Songs list — view kennel songs",       "REGRESSION", "APP",
     "Open a kennel's song list. Verify songs load, play if audio is available, and are sortable."),
    ("TC-160", "Songs management — add song (portal)", "REGRESSION", "APP+PORTAL",
     "In the portal, add a new song to a kennel. In the app, trigger a sync. Verify the new song appears in the kennel song list."),
    ("TC-161", "Settings — distance preference persists","REGRESSION","APP",
     "Change distance preference from km to miles. Force-close and relaunch the app. Verify the preference is still set to miles."),
    ("TC-162", "Offline mode ribbon — appears and clears","REGRESSION","APP",
     "Enable airplane mode. Verify the offline ribbon appears. Disable airplane mode and trigger a sync. Verify the ribbon disappears after successful sync."),
    ("TC-163", "Kennel discovery — new kennel appears","REGRESSION","APP",
     "Create a new active kennel in the portal. In the app, run a discovery/search. Verify the new kennel appears."),
    ("TC-164", "Future runs list — refreshes after portal edit","REGRESSION","APP+PORTAL",
     "Edit a future run in the portal (change the date). In the app, DataChangeService should notify FutureRunListPageController. Verify the updated date appears in the app run list without manual restart."),
    ("TC-165", "KennelsListPage — follow/unfollow",    "REGRESSION", "APP",
     "Follow and unfollow a kennel from the kennels list. Verify the list updates immediately without flicker, sort order is maintained, and the change persists across app restart."),
    ("TC-166", "Check-in sheet — loads with many attendees","REGRESSION","APP",
     "Open the check-in sheet for a run with 50+ attendees. Verify the list loads completely and is scrollable without performance issues."),
    ("TC-167", "Run map — track renders correctly",    "REGRESSION", "APP",
     "Open the map for a past run with a recorded GPS track. Verify the track renders as a continuous line, start/finish markers are shown, and the map pans to fit the track."),
    ("TC-168", "Portal run list — loads and sorts",    "REGRESSION", "PORTAL",
     "Open the portal run list for a kennel. Verify runs load in reverse chronological order, past runs are distinguishable from upcoming runs."),
    ("TC-169", "Portal — stale auth credentials fix",  "REGRESSION", "PORTAL",
     "In the portal, let the session remain idle for an extended period (simulate stale credentials). Verify the auth flow gracefully re-authenticates rather than showing a blank screen or error loop."),
    ("TC-170", "App stability — 30 minute session",    "REGRESSION", "APP",
     "Use the app continuously for 30 minutes: browse runs, open chats, view maps, check admin. Verify no crashes, memory warnings, or unexpected logouts occur throughout the session."),
]

def make_border(style="thin", color=BORDER_COL):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HC3 Test Plan"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # Serial
    ws.column_dimensions["B"].width = 38   # Test Name
    ws.column_dimensions["C"].width = 22   # Category
    ws.column_dimensions["D"].width = 14   # Platform
    ws.column_dimensions["E"].width = 80   # Description
    ws.column_dimensions["F"].width = 16   # Status
    ws.column_dimensions["G"].width = 30   # Notes

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Serial", "Test Name", "Category", "Platform", "Description", "Status", "Notes"]
    hdr_font   = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
    hdr_fill   = PatternFill("solid", fgColor=HEADER_BG)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_border = make_border("medium", "1E3A5F")

    for col, txt in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=txt)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border
    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    thin_border = make_border()
    status_opts  = "Pass / Fail / Blocked / Skip / Not Run"

    prev_cat = None
    for row_idx, (serial, name, cat, platform, desc) in enumerate(TESTS, 2):
        bg_hex = CAT_COLORS.get(cat, "FFFFFF")
        if row_idx % 2 == 0:
            # slightly darken alternate rows within same category
            bg_hex = bg_hex  # keep category colour on even rows
        row_fill = PatternFill("solid", fgColor=bg_hex)

        values = [serial, name, cat, platform, desc, "", ""]
        aligns = [
            Alignment(horizontal="center", vertical="top"),
            Alignment(horizontal="left",   vertical="top", wrap_text=True),
            Alignment(horizontal="center", vertical="top", wrap_text=True),
            Alignment(horizontal="center", vertical="top"),
            Alignment(horizontal="left",   vertical="top", wrap_text=True),
            Alignment(horizontal="center", vertical="top"),
            Alignment(horizontal="left",   vertical="top", wrap_text=True),
        ]
        fonts = [
            Font(name="Calibri", size=10, bold=True),
            Font(name="Calibri", size=10, bold=True),
            Font(name="Calibri", size=10),
            Font(name="Calibri", size=10),
            Font(name="Calibri", size=10),
            Font(name="Calibri", size=10, color="5B5B5B"),
            Font(name="Calibri", size=10, color="5B5B5B"),
        ]
        for col_idx, (val, aln, fnt) in enumerate(zip(values, aligns, fonts), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.alignment = aln
            cell.font      = fnt
            cell.border    = thin_border

        ws.row_dimensions[row_idx].height = max(30, min(120, len(desc) // 3 + 18))

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:G{len(TESTS)+1}"

    # ── Summary tab ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10

    from collections import Counter
    cat_counts = Counter(t[2] for t in TESTS)
    plat_counts = Counter(t[3] for t in TESTS)

    hdr_fill2 = PatternFill("solid", fgColor=HEADER_BG)
    hdr_font2 = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)

    ws2["A1"] = "Harrier Central 3.0 — Test Plan Summary"
    ws2["A1"].font = Font(bold=True, name="Calibri", size=14, color=HEADER_BG)
    ws2["A2"] = f"Total test cases: {len(TESTS)}"
    ws2["A2"].font = Font(bold=True, name="Calibri", size=11)
    ws2["A3"] = ""

    ws2["A4"] = "Category"
    ws2["B4"] = "Count"
    for c in ["A4", "B4"]:
        ws2[c].fill   = hdr_fill2
        ws2[c].font   = hdr_font2
        ws2[c].alignment = Alignment(horizontal="center")

    for i, (cat, count) in enumerate(sorted(cat_counts.items()), 5):
        ws2[f"A{i}"] = cat
        ws2[f"B{i}"] = count
        bg = CAT_COLORS.get(cat, "FFFFFF")
        for c in [f"A{i}", f"B{i}"]:
            ws2[c].fill = PatternFill("solid", fgColor=bg)
            ws2[c].font = Font(name="Calibri", size=10)
            ws2[c].border = make_border()

    row = 5 + len(cat_counts) + 2
    ws2[f"A{row}"] = "Platform"
    ws2[f"B{row}"] = "Count"
    for c in [f"A{row}", f"B{row}"]:
        ws2[c].fill = hdr_fill2
        ws2[c].font = hdr_font2
        ws2[c].alignment = Alignment(horizontal="center")

    for i, (plat, count) in enumerate(sorted(plat_counts.items()), row+1):
        ws2[f"A{i}"] = plat
        ws2[f"B{i}"] = count
        ws2[f"A{i}"].font = Font(name="Calibri", size=10)
        ws2[f"B{i}"].font = Font(name="Calibri", size=10)
        ws2[f"A{i}"].border = make_border()
        ws2[f"B{i}"].border = make_border()

    return wb

if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "HC3_Test_Plan.xlsx")
    wb  = create_workbook()
    wb.save(out)
    print(f"Saved: {out}  ({len(TESTS)} test cases)")
